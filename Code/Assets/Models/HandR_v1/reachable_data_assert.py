import os
import openpyxl

#   读取数据文件
#   插入表格
#   对应行列进行插入操作

work_row = 1  # 当前插入到哪一行
work_number = 12  # 当前录入人员的编号 1-12
work_object = '立方体'  # 立方体 扁平长方体 球体 椭圆体 圆柱 扁平圆柱 圆锥体 半球体 三角柱 六角棱柱

# 读取要插入的数据
with open("D:/VS_workspace/GraspSimulator/Assets/Models/HandR_v1/Results/computed_data1.txt", "r") as file:
    oxy = file.readline().replace('\n', '')
    oxz = file.readline().replace('\n', '')
    oyz = file.readline().replace('\n', '')
    txy = file.readline().replace('\n', '')
    txy1 = file.readline().replace('\n', '')
    txz = file.readline().replace('\n', '')
    txz1 = file.readline().replace('\n', '')
    tyz = file.readline().replace('\n', '')
    tyz1 = file.readline().replace('\n', '')
    ixy = file.readline().replace('\n', '')
    ixy1 = file.readline().replace('\n', '')
    ixz = file.readline().replace('\n', '')
    ixz1 = file.readline().replace('\n', '')
    iyz = file.readline().replace('\n', '')
    iyz1 = file.readline().replace('\n', '')
    mxy = file.readline().replace('\n', '')
    mxy1 = file.readline().replace('\n', '')
    mxz = file.readline().replace('\n', '')
    mxz1 = file.readline().replace('\n', '')
    myz = file.readline().replace('\n', '')
    myz1 = file.readline().replace('\n', '')
    rxy = file.readline().replace('\n', '')
    rxy1 = file.readline().replace('\n', '')
    rxz = file.readline().replace('\n', '')
    rxz1 = file.readline().replace('\n', '')
    ryz = file.readline().replace('\n', '')
    ryz1 = file.readline().replace('\n', '')
    pxy = file.readline().replace('\n', '')
    pxy1 = file.readline().replace('\n', '')
    pxz = file.readline().replace('\n', '')
    pxz1 = file.readline().replace('\n', '')
    pyz = file.readline().replace('\n', '')
    pyz1 = file.readline().replace('\n', '')

path = r"d:\Users\lenovo\Desktop\Grasp Experiment\experiment\Hi5 实验数据\Hi5 实验数据处理"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('Hi5数据统计.xlsx', data_only=True)  # 返回一个workbook数据类型的值

sheet1 = workbook['Sheet1']  # 获取指定sheet

sheet1_row_max = sheet1.max_row  # 获取行数
work_row = sheet1.max_row + 1

sheet1.cell(row=work_row, column=1).value = work_number
sheet1.cell(row=work_row, column=4).value = oxy
sheet1.cell(row=work_row, column=5).value = oxz
sheet1.cell(row=work_row, column=6).value = oyz
sheet1.cell(row=work_row, column=7).value = txy
sheet1.cell(row=work_row, column=8).value = txy1
sheet1.cell(row=work_row, column=9).value = txz
sheet1.cell(row=work_row, column=10).value = txz1
sheet1.cell(row=work_row, column=11).value = tyz
sheet1.cell(row=work_row, column=12).value = tyz1
sheet1.cell(row=work_row, column=13).value = ixy
sheet1.cell(row=work_row, column=14).value = ixy1
sheet1.cell(row=work_row, column=15).value = ixz
sheet1.cell(row=work_row, column=16).value = ixz1
sheet1.cell(row=work_row, column=17).value = iyz
sheet1.cell(row=work_row, column=18).value = iyz1
sheet1.cell(row=work_row, column=19).value = mxy
sheet1.cell(row=work_row, column=20).value = mxy1
sheet1.cell(row=work_row, column=21).value = mxz
sheet1.cell(row=work_row, column=22).value = mxz1
sheet1.cell(row=work_row, column=23).value = myz
sheet1.cell(row=work_row, column=24).value = myz1
sheet1.cell(row=work_row, column=25).value = rxy
sheet1.cell(row=work_row, column=26).value = rxy1
sheet1.cell(row=work_row, column=27).value = rxz
sheet1.cell(row=work_row, column=28).value = rxz1
sheet1.cell(row=work_row, column=29).value = ryz
sheet1.cell(row=work_row, column=30).value = ryz1
sheet1.cell(row=work_row, column=31).value = pxy
sheet1.cell(row=work_row, column=32).value = pxy1
sheet1.cell(row=work_row, column=33).value = pxz
sheet1.cell(row=work_row, column=34).value = pxz1
sheet1.cell(row=work_row, column=35).value = pyz
sheet1.cell(row=work_row, column=36).value = pyz1

workbook.save(r'd:\Users\lenovo\Desktop\Grasp Experiment\experiment\Hi5 实验数据\Hi5 实验数据处理\Hi5数据统计.xlsx')
